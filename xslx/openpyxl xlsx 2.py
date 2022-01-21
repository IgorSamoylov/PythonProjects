import os
import openpyxl

path, subdirs, files = next(os.walk('rogaikopyta_6_4_2'))

result_list = []
for file in files:
    wb = openpyxl.load_workbook(path + '/' + file)
    sh = wb.active
    result_list.append((sh.cell(row=2, column=2).value, str(sh.cell(row=2, column=4).value)))


result_list.sort(key=lambda tup: tup[0])

result_str = ''
for r in result_list:
  result_str += r[0] + ' ' + r[1] + '\n'  

with open('vedomost.txt', 'w', encoding='utf-8') as f:
    f.write(result_str)

